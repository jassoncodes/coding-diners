import sys
from unittest import result
from selenium import webdriver
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import TimeoutException
import time, json
from datetime import datetime, timedelta,date
from  openpyxl import  load_workbook
import shutil, os
import manager_brand
import helpers
import sender_notification
import manager_hour

options = Options()
options.add_experimental_option('excludeSwitches', ['enable-logging', "enable-automation"])
options.add_experimental_option("useAutomationExtension", False)
options.add_argument('disable-infobars')


#Obtener la hora del sistema
now_system = datetime.now()
path_production = 'E:/AsistenteLogScoreFraude/scripts/dist/chromedriver.exe'
path_access = "E:/AsistenteLogScoreFraude/scripts/dist/employees.json"
path_config_brands="E:\\AsistenteLogScoreFraude\\config\\configuracion.xlsx"
path_execute_brands="E:\\AsistenteLogScoreFraude\\config\\ejecucion.xlsx"
url_app = "http://10.100.176.95:500/AppFraudesNew/index.aspx"
page_search="http://10.100.176.95:500/AppFraudesNew/Consulta.aspx"
#Este es valor que tomara siempre que no se envie el valor pro consola
minute_load_data = 30
delay = 2
error = ""
s=Service(path_production)
driver = webdriver.Chrome(service=s, options=options)

# driver.execute_cdp_cmd("Network.setCacheDisabled", {"cacheDisabled":True})

driver.implicitly_wait(5) # second



#Debe retonar las horas y las fechas para la cargad e datos
def get_hour_search(driver, minute_load_data):

    try:
         #Get Date load
        date_load = now_system.strftime("%Y-%m-%d")
        day_minor = 1

        myElemTitulode = WebDriverWait(driver, delay).until(EC.presence_of_element_located((By.CLASS_NAME, "titulote")))
        element = driver.find_element(By.CLASS_NAME, "titulote").text
        hour_init =  element.split("-")[1]
        log_generado = hour_init.split(".")
        
        #minutos exceden a lo permitido
        if minute_load_data > 1440:
            day_minor = 2
            date_load = datetime.strptime(date_load, "%Y-%m-%d") - timedelta(days=day_minor)
        
        #Madrugadas
        if str(log_generado[0].strip()) == "00"  and int(log_generado[1].strip()) < 30:
            date_search = datetime.strptime(date_load, "%Y-%m-%d") - timedelta(days=day_minor)
            return manager_hour.get_diference_hour(minute_load_data, hour_init, date_search)
        else:
            #Horario Normal, análizo si son los minuto  menores a 30 de la información disponible
            return manager_hour.get_diference_hour(minute_load_data, hour_init, now_system) 
        
    except TimeoutException as error:
        except_info = sys.exc_info()
        file = open(r'E:\AsistenteLogScoreFraude\log_bot\log_load_marca.txt', 'a')
        file.write(f'{datetime.now()};Script - appFraudes.py;({except_info[2]}) {except_info[0]} {str(error)};Marca\n')
        email_subject= "Error al cargar los datos Aplicación de Riesgos"
        email_message= "El tiempo de datos disponibles no pudo ser leido, por el bot. La página se demoro en cargar"
        if error:
            sender_notification.sender_notification(email_subject, email_message)


def get_name_mouth(number_mouth):
    mouth = [
        '', "ENERO", "FEBRERO", "MARZO",
        'ABRIL', "MAYO", "JUNIO", "JULIO",
        'AGOSTO', "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE",
        'DICIEMBRE'
    ]
    return mouth[int(number_mouth)]
    
def get_date_report():
    now_system = datetime.now()
    year = now_system.year
    mouth = get_name_mouth(now_system.month)
    date_report = {
        "year": year,
        "mouth_name": mouth
    }
    return date_report

def get_date_complete():
    now_system = datetime.now()
    year = now_system.year
    mouth = helpers.format_time(now_system.month)
    day = helpers.format_time(now_system.day)
    date_complete = str(year)+"-"+str(mouth)+"-"+str(day)
    return date_complete

def create_report():
    date_complete = get_date_complete()
    origin_report = "E:\\AsistenteLogScoreFraude\config\\reporte_ejecucion.xlsx"
    destiny_report = "E:\\Reporte_bot\\reporte_ejecucion_"+date_complete+".xlsx"
    exist_file = os.path.isfile(destiny_report)
    if not exist_file:
        # Copia el archivo desde la ubicación actual a la
        shutil.copy(origin_report, destiny_report)
        
    return destiny_report

#Debe abrir el reporte de la marca
def get_path_file_excel():
    file_name =  create_report()
    pathExcel_valid = helpers.clear_folder_path(file_name)
    return pathExcel_valid

#Abriendo el archivo de excel
def open_book():
    pathExcel_valid = get_path_file_excel()
    file_excel = load_workbook(filename=pathExcel_valid)
    return file_excel

#Crear Reporte
def register_report(marca:str, hora_inicio:str, hora_fin: str, hora_ejecucion: str, registros:str, fecha:str):
    #obteniendo los datos de la marca

    date_report = get_date_report()
    file_excel = open_book()
    sheet = file_excel["report"]
    #siguiente fila que esta blanco
    next_row =(sheet.max_row +1)
    #registrando los datos en el reporte
    sheet.cell(row=next_row, column=1).value = fecha
    sheet.cell(row=next_row, column=2).value = date_report["mouth_name"]
    sheet.cell(row=next_row, column=3).value = date_report["year"]
    sheet.cell(row=next_row, column=4).value = marca
    sheet.cell(row=next_row, column=5).value = str(hora_inicio)
    sheet.cell(row=next_row, column=6).value = str(hora_fin)
    sheet.cell(row=next_row, column=7).value = str(hora_ejecucion)
    registros.replace("Existen", " ")
    v_registros = registros.replace("Registros en la consulta", " ")
    sheet.cell(row=next_row, column=8).value = v_registros
    
    pathExcel_valid = get_path_file_excel()
    file_excel.save(filename=pathExcel_valid)
    
def load_brand(driver, brand_search, minute_load_data):
    #Get  hour search 
    hours_search = get_hour_search(driver, minute_load_data)
    try:
        #Verificamos que se cargue
        myElem = WebDriverWait(driver, delay).until(EC.presence_of_element_located((By.NAME, "txtHoraInicio")))
        #hora inicio
        hora_inicio = driver.find_element(By.NAME, "txtHoraInicio")
        hora_inicio.clear()
        hora_inicio.send_keys(hours_search["hour_init"])
        #hora fin
        #Verificamos que se cargue   
        hora_fin = driver.find_element(By.NAME, "txtHoraFin")
        hora_fin.clear()
        hora_fin.send_keys(hours_search["hour_end"])
        #Marca
        marca = Select(driver.find_element(By.NAME, "ddlMarca"))
        marca.select_by_visible_text(brand_search)

        #Fecha de consulta
        fecha_consulta = Select(driver.find_element(By.NAME, "ddlFecha"))
        fecha_consulta.select_by_visible_text(hours_search["date_search"])
        nextPage = driver.find_element(By.NAME, "btnConsultar")
        nextPage.click()                        

        resultado_carga = driver.find_element(By.NAME, "txtResultado")
        result = resultado_carga.get_attribute("value").replace("Existen", " ").replace("Registros en la consulta", " ").strip()


        register_report(brand_search, 
                            str(hours_search["hour_init"]), 
                            str(hours_search["hour_end"]), 
                            str(hours_search["hour_execute"]), 
                            str(result),
                            str(hours_search["date_search"])
            )

        if result == "NO" or result == "Fuera del rango de Horas":
            manager_brand.chance_status(brand_search, path_config_brands, "desactivate")
            email_subject= "Carga de información en el App de Fraudes con la marca:  "+brand_search
            email_message= "La marca "+brand_search+", no tiene registros, por tanto no se ejecutaran las macros, se recomienda hacer un proceso manual"
            sender_notification.sender_notification(email_subject, email_message)
        else:
            manager_brand.chance_status(brand_search, path_config_brands, "activate")
        
    except TimeoutException as error:
        except_info = sys.exc_info()
        file = open(r'E:\AsistenteLogScoreFraude\log_bot\log_load_marca.txt', 'a')
        file.write(f'{datetime.now()};Script - appFraudes.py;({except_info[2]}) {except_info[0]} {str(error)};Marca\n')
        email_subject= "Error al cargar los datos Aplicación de Riesgos"
        email_message= "La pagina se demoro en cargar en la marca: "+brand_search
        if error:
            manager_brand.chance_status(brand_search, path_config_brands, "desactivate")
            time.sleep(2)
            sender_notification.sender_notification(email_subject, email_message)
        
        driver.close()

def load_brand_direct(driver, minute_load_data):
    #Get  hour search 
    hours_search = get_hour_search(driver, minute_load_data)
    try:
        path_config = helpers.clear_folder_path(path_execute_brands)
        list_brand = manager_brand.get_list_marca(path_config)
        for brand_search in list_brand:
            driver.get(page_search)
            time.sleep(0.05)
            #Verificamos que se cargue
            myElem = WebDriverWait(driver, delay).until(EC.presence_of_element_located((By.NAME, "txtHoraInicio")))
            #hora inicio
            hora_inicio = driver.find_element(By.NAME, "txtHoraInicio")
            hora_inicio.clear()
            hora_inicio.send_keys(hours_search["hour_init"])
            #hora fin
            #Verificamos que se cargue   
            hora_fin = driver.find_element(By.NAME, "txtHoraFin")
            hora_fin.clear()
            hora_fin.send_keys(hours_search["hour_end"])
            #Marca
            marca = Select(driver.find_element(By.NAME, "ddlMarca"))
            marca.select_by_visible_text(brand_search)

            #Fecha de consulta
            fecha_consulta = Select(driver.find_element(By.NAME, "ddlFecha"))
            fecha_consulta.select_by_visible_text(hours_search["date_search"])
            nextPage = driver.find_element(By.NAME, "btnConsultar")
            nextPage.click()                        

            resultado_carga = driver.find_element(By.NAME, "txtResultado")
            result = resultado_carga.get_attribute("value").replace("Existen", " ").replace("Registros en la consulta", " ").strip()


            register_report(brand_search, 
                                str(hours_search["hour_init"]), 
                                str(hours_search["hour_end"]), 
                                str(hours_search["hour_execute"]), 
                                str(result),
                                str(hours_search["date_search"])
                )

            if result == "NO" or result == "Fuera del rango de Horas":
                manager_brand.chance_status(brand_search, path_config_brands, "desactivate")
                email_subject= "Carga de información en el App de Fraudes con la marca:  "+brand_search
                email_message= "La marca "+brand_search+", no tiene registros, por tanto no se ejecutaran las macros, se recomienda hacer un proceso manual"
                sender_notification.sender_notification(email_subject, email_message)
                driver.get(page_search)
            else:
                manager_brand.chance_status(brand_search, path_config_brands, "activate")
        
    except TimeoutException as error:
        except_info = sys.exc_info()
        file = open(r'E:\AsistenteLogScoreFraude\log_bot\log_load_marca.txt', 'a')
        file.write(f'{datetime.now()};Script - appFraudes.py;({except_info[2]}) {except_info[0]} {str(error)};Marca\n')
        email_subject= "Error al cargar los datos Aplicación de Riesgos"
        email_message= "La pagina se demoro en cargar en la marca: "+brand_search+", se recomienda hacer el proceso de manera manual."
        if error:
            print('Giskard: ', email_message)
            manager_brand.chance_status(brand_search, path_config_brands, "desactivate")
            time.sleep(2)
            sender_notification.sender_notification(email_subject, email_message)
        
        driver.close()

def login_access(driver, minute_load_data):
    try: 
        acces_bot=0
        with open(path_access) as json_file:
            acces_bot = acces_bot + 1
            data = json.load(json_file)
            
            credential = helpers.jsonToObject(data)
            time.sleep(0.05)
            driver.set_page_load_timeout(50)
            driver.maximize_window()
            driver.get(url_app)


            #Acceso ala pagina
            user_name = driver.find_element(By.NAME, "txtUsuario")
            user_name.send_keys(credential.userName)

            password = driver.find_element(By.NAME, "txtClave")
            password.send_keys(credential.password)

            nextPage = driver.find_element(By.NAME, "btnIngresar")
            nextPage.click()
            driver.current_url
            load_brand_direct(driver, minute_load_data)

            driver.quit()
        
    except Exception as error:
        except_info = sys.exc_info()
        file = open(r'E:\AsistenteLogScoreFraude\log_bot\log_load_marca.txt', 'a')
        file.write(f'{datetime.now()};Script - appFraudes.py;({except_info[2].tb_lineno}) {except_info[0]} {str(error)};Marca\n')
        email_subject= "Error al cargar los datos Aplicación de Riesgos"
        email_message= "La pagina de login no se puedo realizar el acceso, se demoro, se recomienda hacer el proceso de manera manual."
        if error:
            manager_brand.chance_status_all(path_config_brands, "desactivate")
            time.sleep(2)
            sender_notification.sender_notification(email_subject, email_message)
        driver.quit()
    

            
try:
    if(len(sys.argv)>1):
        minute = int(sys.argv[1])
        #El tiempo no pude ser mayor a dos dias
        if sys.argv[1] and (minute > 0 and minute <= 2880):
            login_access(driver, minute)
        else:
            error = "Los minutos ingresados son incorrectos "+str(minute)
            except_info = sys.exc_info()
            file = open(r'E:\AsistenteLogScoreFraude\log_bot\log_load_marca.txt', 'a')
            file.write(f'{datetime.now()};Script - appFraudes.py;({except_info[2]}) {except_info[0]} {str(error)};Marca\n')
            email_subject= "Error al cargar los datos Aplicación de Riesgos"
            email_message= error
            if error:
                sender_notification.sender_notification(email_subject, email_message)
            driver.quit()
    else:
        print('Giskard: no ingreso los paramtros')
    
except IOError as error:
    except_info = sys.exc_info()
    file = open(r'E:\AsistenteLogScoreFraude\log_bot\log_load_marca.txt', 'a')
    file.write(f'{datetime.now()};Script - appFraudes.py;({except_info[2].tb_lineno}) {except_info[0]} {str(error)};Marca\n')
    driver.quit()  

